"""Excel (.xlsx / .xls) parser for security questionnaires.

Opens workbooks with openpyxl in read-only mode, scans each sheet for
table-like structures, and extracts question/answer pairs using the same
heuristic patterns as the DOCX and CSV parsers.  Location info includes
sheet name, row number, and column indices so the generator can write
answers back into the correct cells.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from app.services.parsers.types import ExtractedItem, ParseOptions, ParseResult

# ---------------------------------------------------------------------------
# Heuristic helpers (mirrored from heuristic.py to avoid circular imports)
# ---------------------------------------------------------------------------

QUESTION_PATTERNS = [
    re.compile(r"\?\s*$"),
    re.compile(r"^\d+[\.\)]\s+.{10,}"),
    re.compile(r"^[a-zA-Z][\.\)]\s+.{10,}"),
    re.compile(
        r"^(please\s+)?(describe|explain|provide|list|detail|specify|state|outline|indicate|confirm|identify)",
        re.IGNORECASE,
    ),
]

NON_QUESTION_PATTERNS = [
    re.compile(r"\bplease\s+provide\s+detailed\s+responses?\s+to\s+(each|all)\s+questions?\s+below\b", re.IGNORECASE),
    re.compile(r"\bdetailed\s+responses?\s+required\b", re.IGNORECASE),
    re.compile(r"\bdetailed\s+assessment\s+with\s+verbose\s+questions\b", re.IGNORECASE),
]

LABEL_PATTERNS = [
    re.compile(
        r"^(company\s+name|organization|contact|address|phone|email|website|date|name|title|role|department)",
        re.IGNORECASE,
    ),
    re.compile(
        r"^(policy|procedure|description|overview|summary|response|answer|comments?|notes?)\s*:?\s*$",
        re.IGNORECASE,
    ),
]

QUESTION_HEADER_PATTERNS = [
    re.compile(r"\bquestion(s)?\b", re.IGNORECASE),
    re.compile(r"\b(prompt|requirement|control|field)\b", re.IGNORECASE),
]

ANSWER_HEADER_PATTERNS = [
    re.compile(r"\banswer(s)?\b", re.IGNORECASE),
    re.compile(r"\bresponse(s)?\b", re.IGNORECASE),
    re.compile(r"\b(comment|comments|notes?)\b", re.IGNORECASE),
]


def _is_question(text: str, min_length: int = 10) -> bool:
    text = text.strip()
    if len(text) < min_length:
        return False
    for pat in NON_QUESTION_PATTERNS:
        if pat.search(text):
            return False
    for pat in QUESTION_PATTERNS:
        if pat.search(text):
            return True
    for pat in LABEL_PATTERNS:
        if pat.search(text):
            return True
    return False


def _matches_any(text: str, patterns: list[re.Pattern]) -> bool:
    for pat in patterns:
        if pat.search(text):
            return True
    return False


# ---------------------------------------------------------------------------
# Column auto-detection (same algorithm as _infer_table_mapping)
# ---------------------------------------------------------------------------

def _infer_columns(
    rows: list[list[str]],
    options: ParseOptions,
) -> tuple[int, int, int]:
    """Return (question_col, answer_col, header_rows) for a sheet region."""

    q_col = options.question_column_index
    a_col = options.answer_column_index
    header_rows = options.header_rows

    if not rows or not options.auto_detect_columns:
        return q_col, a_col, header_rows

    first_row = [cell.strip() for cell in rows[0]]
    if first_row:
        q_headers = [i for i, t in enumerate(first_row) if _matches_any(t, QUESTION_HEADER_PATTERNS)]
        a_headers = [i for i, t in enumerate(first_row) if _matches_any(t, ANSWER_HEADER_PATTERNS)]
        if q_headers and a_headers:
            q_col = q_headers[0]
            right = [i for i in a_headers if i > q_col]
            a_col = right[0] if right else a_headers[0]
            return q_col, a_col, max(header_rows, 1)

    if len(first_row) <= 2:
        return q_col, a_col, header_rows

    scores: dict[tuple[int, int], int] = {}
    for row in rows[header_rows:]:
        texts = [t.strip() for t in row]
        for ci, qt in enumerate(texts[:-1]):
            if not qt or not _is_question(qt, options.min_question_length):
                continue
            for ai in range(ci + 1, len(texts)):
                if len(texts[ai]) >= 3:
                    continue
                score = 3
                if ai == ci + 1:
                    score += 2
                if all(len(texts[p].strip()) < options.min_question_length for p in range(ci)):
                    score += 1
                if any(texts[m].strip() for m in range(ci + 1, ai)):
                    score -= 2
                scores[(ci, ai)] = scores.get((ci, ai), 0) + score

    if scores:
        q_col, a_col = max(scores.items(), key=lambda x: (x[1], -x[0][0], -x[0][1]))[0]

    return q_col, a_col, header_rows


# ---------------------------------------------------------------------------
# Confidence scoring
# ---------------------------------------------------------------------------

def _score_confidence(stats: dict[str, Any], items: list[ExtractedItem]) -> float:
    total_rows = stats.get("rows_scanned", 0)
    if total_rows == 0 or not items:
        return 0.0
    ratio = len(items) / max(total_rows, 1)
    if ratio > 0.8:
        return 0.95
    if ratio > 0.4:
        return 0.85
    if ratio > 0.1:
        return 0.7
    return 0.5


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_excel(file_path: Path, options: ParseOptions | None = None) -> ParseResult:
    """Parse an Excel workbook and return question/answer items.

    Each item carries enough location info (sheet_name, row, q_col_idx,
    a_col_idx) for the generator to write answers back into the original
    workbook.
    """

    options = options or ParseOptions()
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    all_items: list[ExtractedItem] = []
    total_rows = 0
    total_sheets = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            total_sheets += 1

            # Materialise rows into plain strings
            raw_rows: list[list[str]] = []
            for row in ws.iter_rows():
                raw_rows.append([str(cell.value).strip() if cell.value is not None else "" for cell in row])

            if not raw_rows:
                continue

            total_rows += len(raw_rows)
            q_col, a_col, header_rows = _infer_columns(raw_rows, options)

            for r_idx, row_values in enumerate(raw_rows):
                if r_idx < header_rows:
                    continue

                max_idx = max(q_col, a_col)
                if len(row_values) <= max_idx:
                    continue
                if not options.allow_additional_columns and len(row_values) > max_idx + 1:
                    continue

                q_text = row_values[q_col].strip()
                if not q_text or len(q_text) < options.min_question_length:
                    continue

                if _is_question(q_text, options.min_question_length):
                    # Excel rows in openpyxl are 1-based; we store both
                    # the 0-based list index and the 1-based Excel row for
                    # the generator (which opens the workbook in write mode
                    # using 1-based indexing).
                    excel_row = r_idx + 1  # 1-based
                    all_items.append(
                        ExtractedItem(
                            question_text=q_text,
                            item_type="excel_cell",
                            location={
                                "sheet_name": sheet_name,
                                "row_idx": r_idx,
                                "excel_row": excel_row,
                                "q_col_idx": q_col,
                                "a_col_idx": a_col,
                                "question_col_idx": q_col,
                                "answer_col_idx": a_col,
                                "column_count": len(row_values),
                                "header_rows": header_rows,
                            },
                            source_block_id=f"xlsx-{sheet_name}-r{r_idx}-q{q_col}-a{a_col}",
                            confidence=0.88,
                            parser_strategy="excel_heuristic",
                            raw_text=q_text,
                        )
                    )
    finally:
        wb.close()

    stats = {
        "sheets_scanned": total_sheets,
        "rows_scanned": total_rows,
        "excel_items": len(all_items),
        "items_total": len(all_items),
    }

    confidence = _score_confidence(stats, all_items)
    fallback_recommended = confidence < 0.4 and len(all_items) < 3
    fallback_reason = "Very few questions detected in workbook" if fallback_recommended else None

    return ParseResult(
        items=all_items,
        confidence=confidence,
        stats=stats,
        profile_name=options.profile_name,
        parser_strategy="excel_heuristic",
        fallback_recommended=fallback_recommended,
        fallback_reason=fallback_reason,
    )
