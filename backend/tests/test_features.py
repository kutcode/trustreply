"""Tests for F2 (contradiction detection), F3 (SME routing), F5 (XLSX).

These are the three features shipped after the initial test suite was
written, so they had zero coverage until now. Each test exercises the
piece most likely to regress: persistence for F2, the side effect on
resolution for F3, and the file round-trip for F5.
"""

import openpyxl
import pytest_asyncio

from app.config import settings
from app.models import DuplicateReview, FlaggedQuestion, ProcessingJob, QAPair
from app.services.generator import generate_filled_xlsx
from app.services.parsers.excel_parser import parse_excel
from app.services.parsers.types import ExtractedItem


# ---------------------------------------------------------------------------
# F2 - Contradiction detection
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def kb_pair_factory(db_session):
    """Create QAPair rows without hand-rolling the full init each time."""
    async def _make(question: str, answer: str, category: str = "General") -> QAPair:
        pair = QAPair(category=category, question=question, answer=answer)
        db_session.add(pair)
        await db_session.commit()
        await db_session.refresh(pair)
        return pair
    return _make


async def test_contradictions_count_endpoint_reflects_flagged_pairs(client, db_session, kb_pair_factory):
    """GET /api/qa/contradictions/count returns the number of pairs where
    contradicts=True. Without this feature, conflicting KB answers would
    pile up silently; the count badge is how reviewers notice."""
    a1 = await kb_pair_factory("Do you encrypt data at rest?", "Yes, AES-256.")
    a2 = await kb_pair_factory("Do you encrypt data at rest?", "No, we do not.")
    b1 = await kb_pair_factory("What is your MFA policy?", "MFA required for admin.")
    b2 = await kb_pair_factory("What is your MFA policy?", "MFA required for admin accounts.")

    # Contradicting pair
    db_session.add(DuplicateReview(
        entry_a_id=a1.id, entry_b_id=a2.id,
        similarity_score=0.91, classification="probably_same",
        contradicts=True, status="pending",
    ))
    # Non-contradicting near-duplicate (same intent, just reworded)
    db_session.add(DuplicateReview(
        entry_a_id=b1.id, entry_b_id=b2.id,
        similarity_score=0.97, classification="definitely_same",
        contradicts=False, status="pending",
    ))
    await db_session.commit()

    res = await client.get("/api/qa/contradictions/count")
    assert res.status_code == 200
    assert res.json() == {"count": 1}


# ---------------------------------------------------------------------------
# F3 - SME routing
# ---------------------------------------------------------------------------


async def test_resolve_flagged_assigns_sme_email_by_category(client, db_session):
    """When SME routing is enabled and the resolved category is in the
    map, the flagged row gets tagged with the SME's email so reviewers
    can see at a glance who the question was routed to."""
    settings.sme_routing_enabled = True
    settings.category_sme_map = {
        "Security": "security-team@example.com",
        "Privacy": "privacy-team@example.com",
    }

    job = ProcessingJob(
        original_filename="q.docx",
        stored_filename="q_stored.docx",
        status="done",
    )
    db_session.add(job)
    await db_session.commit()
    await db_session.refresh(job)

    flag = FlaggedQuestion(
        job_id=job.id,
        extracted_question="Describe how you encrypt data at rest.",
        resolved=False,
    )
    db_session.add(flag)
    await db_session.commit()
    await db_session.refresh(flag)

    res = await client.post(
        f"/api/flagged/{flag.id}/resolve",
        json={
            "answer": "We use AES-256 at rest.",
            "add_to_knowledge_base": True,
            "category": "Security",
        },
    )
    assert res.status_code == 200

    await db_session.refresh(flag)
    assert flag.resolved is True
    assert flag.assigned_to == "security-team@example.com"


async def test_resolve_flagged_skips_assignment_when_category_not_mapped(client, db_session):
    """Categories outside the SME map must leave assigned_to unset.
    Regression guard: an earlier version fell back to a default email."""
    settings.sme_routing_enabled = True
    settings.category_sme_map = {"Security": "security-team@example.com"}

    job = ProcessingJob(
        original_filename="q.docx",
        stored_filename="q_stored.docx",
        status="done",
    )
    db_session.add(job)
    await db_session.commit()
    await db_session.refresh(job)

    flag = FlaggedQuestion(
        job_id=job.id,
        extracted_question="What is your data retention period?",
        resolved=False,
    )
    db_session.add(flag)
    await db_session.commit()
    await db_session.refresh(flag)

    res = await client.post(
        f"/api/flagged/{flag.id}/resolve",
        json={
            "answer": "Seven years.",
            "add_to_knowledge_base": True,
            "category": "Compliance",  # not in the map
        },
    )
    assert res.status_code == 200

    await db_session.refresh(flag)
    assert flag.resolved is True
    assert flag.assigned_to is None


# ---------------------------------------------------------------------------
# F5 - XLSX round-trip
# ---------------------------------------------------------------------------


def _make_xlsx(path, rows):
    """Create a minimal .xlsx file with a header row plus the given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return path


def test_parse_excel_extracts_question_rows(tmp_path):
    """The Excel parser must recognise the Question / Answer columns
    automatically and emit an ExtractedItem per question row with enough
    location info for the generator to round-trip."""
    path = _make_xlsx(
        tmp_path / "security.xlsx",
        [
            ["Do you encrypt data at rest?", ""],
            ["Describe your incident response process.", ""],
            ["", ""],  # blank row should be skipped
        ],
    )

    result = parse_excel(path)
    assert len(result.items) == 2

    questions = {item.question_text for item in result.items}
    assert questions == {
        "Do you encrypt data at rest?",
        "Describe your incident response process.",
    }

    first = result.items[0]
    assert first.item_type == "excel_cell"
    assert first.location["sheet_name"] == "Sheet"
    assert first.location["q_col_idx"] == 0
    assert first.location["a_col_idx"] == 1
    assert first.location["excel_row"] == 2  # row 1 is the header


def test_generate_filled_xlsx_writes_answers_into_original_cells(tmp_path):
    """Round-trip: parse -> fill -> reopen must produce a workbook whose
    answer cells contain the values we supplied. The generator writes into
    the original file so styles / dropdowns survive; we verify the answer
    text landed in the right cell by reopening the output."""
    source = _make_xlsx(
        tmp_path / "in.xlsx",
        [
            ["Do you encrypt data at rest?", ""],
            ["Describe your incident response process.", ""],
        ],
    )
    output = tmp_path / "out.xlsx"

    parsed = parse_excel(source)
    assert len(parsed.items) == 2

    answers = {
        "Do you encrypt data at rest?": "Yes, AES-256 via KMS.",
        "Describe your incident response process.": "Documented, tabletop-tested quarterly.",
    }
    filled: list[ExtractedItem] = []
    for item in parsed.items:
        filled.append(ExtractedItem(
            question_text=item.question_text,
            answer_text=answers[item.question_text],
            item_type=item.item_type,
            location=item.location,
            source_block_id=item.source_block_id,
            confidence=item.confidence,
            parser_strategy=item.parser_strategy,
            raw_text=item.raw_text,
        ))

    generate_filled_xlsx(source, output, filled)

    wb = openpyxl.load_workbook(str(output), data_only=True)
    ws = wb.active
    # Row 2 is first question, row 3 is second. Column B (index 2) is the answer.
    assert ws.cell(row=2, column=2).value == "Yes, AES-256 via KMS."
    assert ws.cell(row=3, column=2).value == "Documented, tabletop-tested quarterly."
    # Question column must be untouched.
    assert ws.cell(row=2, column=1).value == "Do you encrypt data at rest?"
    wb.close()


def test_generate_filled_xlsx_marks_unresolved_answers(tmp_path):
    """Answers with answer_text=None should land as a review placeholder
    with an italic/red font cue. This is what reviewers actually see when
    scanning the spreadsheet for gaps."""
    source = _make_xlsx(tmp_path / "in.xlsx", [["What is your data retention period?", ""]])
    output = tmp_path / "out.xlsx"

    parsed = parse_excel(source)
    assert len(parsed.items) == 1
    item = parsed.items[0]

    unresolved = ExtractedItem(
        question_text=item.question_text,
        answer_text=None,  # unresolved on purpose
        item_type=item.item_type,
        location=item.location,
        source_block_id=item.source_block_id,
        confidence=item.confidence,
        parser_strategy=item.parser_strategy,
        raw_text=item.raw_text,
    )

    generate_filled_xlsx(source, output, [unresolved])

    wb = openpyxl.load_workbook(str(output))
    ws = wb.active
    answer_cell = ws.cell(row=2, column=2)
    assert answer_cell.value  # placeholder text was written
    assert answer_cell.font.italic is True
    assert answer_cell.font.color is not None
    assert str(answer_cell.font.color.rgb).endswith("CC0000")
    wb.close()
